from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# === Constants ===
TAB_COLORS = ['FF9999', '99CCFF', 'CCFFCC', 'FFFF99', 'FFCCFF']
HEADER_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# === Sample Data Generator ===
def generate_sample_data(headers, row_index=1):
    return [f"Sample {header} {row_index}" for header in headers]

# === Header Extraction ===
def extract_structured_headers(field_groups):
    top_headers, sub_headers = [], []
    for group, fields in field_groups.items():
        keys = list(fields.keys())
        top_headers.extend([group] * len(keys))
        sub_headers.extend(keys)
    return top_headers, sub_headers

# === Fill Header Rows (with merging and styles) ===
def fill_headers(ws, field_groups):
    top_headers, sub_headers = extract_structured_headers(field_groups)
    ws.append(top_headers)
    ws.append(sub_headers)

    col_count = len(sub_headers)

    # Merge cells for each group in row 1
    group_start = 0
    for i in range(col_count + 1):  # Include sentinel
        if i == col_count or top_headers[i] != top_headers[group_start]:
            if i - group_start > 1:
                ws.merge_cells(start_row=1, start_column=group_start + 1,
                               end_row=1, end_column=i)
            group_start = i

    # Style header rows
    for row in ws.iter_rows(min_row=1, max_row=2, max_col=col_count):
        for cell in row:
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN

    return sub_headers

# === Fill Sample Data Rows ===
def fill_data(ws, headers, row_count=3):
    for i in range(1, row_count + 1):
        ws.append(generate_sample_data(headers, i))

# === Auto Fit Column Widths ===
def autofit_columns(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

# === Main Workbook Creation ===
def create_excel_from_export_fields(export_fields, filename):
    wb = Workbook()
    wb.remove(wb.active)

    for idx, (sheet_name, field_groups) in enumerate(export_fields.items()):
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = TAB_COLORS[idx % len(TAB_COLORS)]

        headers = fill_headers(ws, field_groups)
        fill_data(ws, headers)
        autofit_columns(ws)

    wb.save(filename)

# === Export Fields ===
export_fields = {
    "Summary": {
        "Appraisal Year": {
            "Username": "appraisee.username",
            "Name": "appraisee.full_name",
            "Branch": "appraisee.detail.branch.name",
            "Job Title": "appraisee.detail.job_title.title",
            "Level": "appraisee.detail.employment_level.title",
            "Appraisal Period From": "annual_appraisal.start_date",
            "Appraisal Period To": "annual_appraisal.end_date",
            "Appraiser Name": "appraiser_full_name",
            "Appraiser Username": "appraiser_username",
            "Zone List": "zone_name",
            "Ethics and Integrity Declaration": "ethics_and_integrity_declaration",
            "Ethics and Integrity Declaration Comment": "ethics_and_integrity_comment",
            "Reviewer Name": "reviewer_full_name",
            "Reviewer Username": "reviewer_username",
            "Reviewers Endorsement (Comments)": "reviewer_comment",
            "Appraisee's Acknowledgment Comment": "appraisee_acknowledgment_comment",
            "Revised Zone List": "revised_zone_list",
            "Additional Notes": "additional_notes",
            "Final Zone List": "final_zone_list"
        }
    },
    "Detailed BCD Indicators": {
        "Appraisal Year": {
            "Appraiser Name": "appraiser_full_name",
            "Appraiser Username": "appraiser_username",
        },
        "Behavioral Indicators": {
            "ownership & Initiative": "",
            "Collaboration & Communication": "",
            "Customer & Stakeholder Experience": "",
        },
        "Compliance": {
            "Regulatory & Policy Adherence": "",
            "Job Knowledge & Application": "",
            "Risk Awareness & Accountability": "",
        },
        "Delivery": {
            "Achievement of Assigned Goals & Role-Specific Deliverables": "",
            "Quality of Work & Problem-Solving": "",
            "Process & Operational Efficiency": "",
        },
        "Final": {
            "Final Rating": "",
            "Zone list": "",
            "Final Zone list after revied from HR Committee": ""
        }
    },
    "Performance Gap Indicators": {
        "Appraisal Year": {
            "Username": "appraiser_username",
            "Name": "appraiser_full_name",
            "Branch": "appraisee.detail.branch.name",
            "Job Title": "appraisee.detail.job_title.title",
            "Level": "appraisee.detail.employment_level.title",
        },
        "Performance Gap Indicators": {
            "Performance Area": "",
            "Examples": "",
            "Performance Gap Observed": "",
        }
    },
    "Action Plan": {
        "Appraisal Year": {
            "Username": "appraiser_username",
            "Name": "appraiser_full_name",
            "Branch": "appraisee.detail.branch.name",
            "Job Title": "appraisee.detail.job_title.title",
            "Level": "appraisee.detail.employment_level.title",
        },
        "Recommended Action Plan for Improvement & Support Required": {
            "Action Plan for Improvement": "",
            "Support/Resources Required": ""
        }
    },
}

# === Run Export ===
create_excel_from_export_fields(export_fields, "appraisal_export_optimized_clean.xlsx")
