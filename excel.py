from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.comments import Comment

wb = Workbook()
ws = wb.active
ws.title = "Dynamic Merge Example"

headers = [
    "Name", "Employee Code", "Username", "Department", "Branch", "Province",
    "Adjustment Requested Date/Time", "Adjustment Sent For Date/Time",
    "Number of time adjusted (per Day)", "Attendance Adjustment Category", "Remarks"
]

for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

data = [
    ["John Doe", "E001", "johndoe", "HR", "Main Branch", "California",
     "2024-11-01 08:00", "2024-11-01 10:00", 2, "Late Arrival", "Traffic"],
    ["Jane Smith", "E002", "janesmith", "IT", "Branch A", "California",
     "2024-11-01 08:00", "2024-11-01 11:00", 2, "Late Arrival", "Meeting"],
    ["Alice Brown", "E003", "aliceb", "Finance", "Branch B", "New York",
     "2024-11-02 09:00", "2024-11-02 11:00", 1, "Early Leave", "Doctor"],
    ["Alice Brown", "E003", "aliceb", "Finance", "Branch B", "New York",
     "2024-11-02 09:00", "2024-11-02 11:00", 1, "Early Leave", "Doctor"],
    ["Alice Brown", "E003", "aliceb", "Finance", "Branch B", "New York",
     "2024-11-03 09:00", "2024-11-03 11:00", 1, "Early Leave", "Doctor"],
    ["Alice Brown", "E003", "aliceb", "Finance", "Branch B", "New York",
     "2024-11-03 09:00", "2024-11-03 11:00", 1, "Early Leave", "Doctor"],
    ["Alice Brown", "E003", "aliceb", "Finance", "Branch B", "New York",
     "2024-11-03 09:00", "2024-11-03 11:00", 1, "Early Leave", "Doctor"],
    ["Alice Brown", "E003", "aliceb", "Finance", "Branch B", "New York",
     "2024-11-04 09:00", "2024-11-04 11:00", 1, "Early Leave", "Doctor"],
    ["Alice Brown", "E003", "aliceb", "Finance", "Branch B", "New York",
     "2024-11-05 09:00", "2024-11-05 11:00", 1, "Early Leave", "Doctor"],
    ["Salina", "E003", "aliceb", "Finance", "Branch B", "New York",
    "2024-11-05 09:00", "2024-12-05 11:00", 1, "Early Leave", "Doctor"],
    ["Alice Brown", "E003", "aliceb", "Finance", "Branch B", "New York",
    "2024-11-03 09:00", "2024-11-03 11:00", 1, "Early Leave", "Doctor"],
    ["John Doe", "E001", "johndoe", "HR", "Main Branch", "California",
     "2024-11-01 08:00", "2024-11-01 10:00", 2, "Late Arrival", "Traffic"],
    ["Jane Smith", "E002", "janesmith", "IT", "Branch A", "California",
     "2024-11-01 08:00", "2024-11-01 11:00", 2, "Late Arrival", "Meeting"],
]
data = sorted(data, key=lambda x: (x[0], x[1], x[6]))

for row_num, row_data in enumerate(data, start=2):
    for col_num, value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_num, value=value)

column_to_check = 7 
merge_column = 9  
start_row = 2 
end_row = len(data) + 1

current_value = None
merge_start = start_row

for row in range(start_row, end_row + 1):
    cell_value = ws.cell(row=row, column=column_to_check).value

    if cell_value != current_value:  
        if merge_start < row - 1:
            ws.merge_cells(
                start_row=merge_start, start_column=merge_column,
                end_row=row - 1, end_column=merge_column
            )
            merged_cell = ws.cell(row=merge_start, column=merge_column)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")

        current_value = cell_value
        merge_start = row

comment = Comment('This is the comment text', 'Comment Author')
ws["A1"].comment = comment

notes = [
    {"cell": "B1", "text": "Keep up the good work!", "author": "Manager"},
    {"cell": "C1", "text": "Schedule additional training sessions.", "author": "Manager"},
    {"cell": "D1", "text": "Congratulations on the promotion!", "author": "HR Team"},
]

# Add notes to the specified cells
for note in notes:
    cell = note["cell"]
    comment = Comment(note["text"], note["author"])
    ws[cell].comment = comment


wb.save("dynamic_merge_example.xlsx")
